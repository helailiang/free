#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
newpre.md 角分辨率 + Y100SC 连续步进测试（命令行版）。

不修改 angle_resolution_test_app.py；本脚本可独立运行。

用法（在「码盘补偿」目录）:
  python newpre_resolution_cli_test.py --help

示例:
  python newpre_resolution_cli_test.py --radar-ip 192.168.1.111 --com COM6 --steps 50 --dir +

依赖: pyserial（与 y100sc_client、h1_radar_reader 一致）。
可选导出: pip install openpyxl 后可写 .xlsx；否则使用 --out result.csv。

图形界面（PySide6）见同目录: newpre_resolution_gui_test.py
"""
from __future__ import annotations

import argparse
import csv
import statistics
import sys
import time
from pathlib import Path
from typing import Any

from h1_radar_reader import H1CalibrationRadar
from y100sc_client import Axis, Sign, Y100SCClient, Y100SCError, Y100SCSerialConfig


def index_center_from_run(run: list[dict[str, Any]]) -> int:
    i0 = int(run[0]["index"])
    i1 = int(run[-1]["index"])
    return (i0 + i1) // 2


def find_contiguous_runs(
    all_results: list[dict[str, Any]], min_mm: int, max_mm: int
) -> list[list[dict[str, Any]]]:
    runs: list[list[dict[str, Any]]] = []
    cur: list[dict[str, Any]] = []
    for p in all_results:
        d = int(p["measured_distance"])
        ok = min_mm < d < max_mm
        if not ok:
            if len(cur) >= 2:
                runs.append(cur)
            cur = []
            continue
        idx = int(p["index"])
        if not cur:
            cur = [p]
        elif idx == int(cur[-1]["index"]) + 1:
            cur.append(p)
        else:
            if len(cur) >= 2:
                runs.append(cur)
            cur = [p]
    if len(cur) >= 2:
        runs.append(cur)
    return runs


def find_longest_run_for_calibration(
    all_results: list[dict[str, Any]], min_mm: int, max_mm: int
) -> tuple[list[dict[str, Any]] | None, str | None]:
    runs = find_contiguous_runs(all_results, min_mm, max_mm)
    if not runs:
        return None, f"在距离 ({min_mm},{max_mm}) mm 内未找到≥2 点的连续目标段"
    best = max(runs, key=len)
    return best, None


def pick_run_near_theory(
    all_results: list[dict[str, Any]],
    min_mm: int,
    max_mm: int,
    theory_center: int,
    half_window: int,
) -> tuple[list[dict[str, Any]] | None, str | None]:
    lo = theory_center - half_window
    hi = theory_center + half_window
    runs = find_contiguous_runs(all_results, min_mm, max_mm)
    candidates: list[tuple[int, list[dict[str, Any]]]] = []
    for run in runs:
        i0 = int(run[0]["index"])
        i1 = int(run[-1]["index"])
        if i1 < lo or i0 > hi:
            continue
        c = index_center_from_run(run)
        candidates.append((abs(c - theory_center), run))
    if not candidates:
        return None, f"理论 index≈{theory_center} ±{half_window} 内无有效连续目标"
    candidates.sort(key=lambda t: t[0])
    return candidates[0][1], None


def snapshot_from_run(run: list[dict[str, Any]], angular_res_deg: float) -> dict[str, Any]:
    cidx = index_center_from_run(run)
    mid_pt = min(run, key=lambda p: abs(int(p["index"]) - cidx))
    i0 = int(run[0]["index"])
    i1 = int(run[-1]["index"])
    return {
        "center_index": cidx,
        "angle_deg": float(mid_pt["angle_deg"]),
        "distance_mm": int(mid_pt["measured_distance"]),
        "intensity": int(mid_pt["reflectivity"]),
        "index_start": i0,
        "index_end": i1,
        "run_len": len(run),
        "angular_res_deg": float(angular_res_deg),
    }


def lookup_point_by_index(all_results: list[dict[str, Any]], idx: int) -> dict[str, Any] | None:
    """在当帧 all_results 中按 index 精确取点（与 h1_radar_reader 解析字段一致）。"""
    for p in all_results:
        if int(p["index"]) == int(idx):
            return p
    return None


def build_step_row(
    *,
    step: int,
    rotate_label: str,
    rotate_deg: float,
    target_index: int,
    index_start: int | str | None,
    index_end: int | str | None,
    delta_index: int,
    measured_angle_deg: float,
    theory_angle_deg: float,
    error_deg: float,
    timestamp: str,
    anomalies: str,
    raw: dict[str, Any],
    # H2 等 GUI：距离窗内多次采集、每帧取反射率最高 index 后的中位数，及是否与 target_index 相同（是/否）。
    probe_median_index: int | str | None = None,
    probe_match_target: str | None = None,
    # 各次窗内最高反射 index 序列（逗号分隔）、其中出现次数的全局峰值、target_index − 中位索引。
    probe_indices_sequence: str | None = None,
    # 上述序列是否包含 target_index（是/否），以及 target_index 在该序列中的出现次数（H2 连续测试用）。
    probe_sequence_contains_target: str | None = None,
    probe_sequence_target_hit_count: int | str | None = None,
    probe_max_hit_count: int | str | None = None,
    target_minus_probe_median: int | str | None = None,
    # H2：相对「反射中位 index」左右各偏 neighbor_index_delta 的展示 index；邻距为各帧主峰±Δ 的均值（mm）；
    # neighbor_left_std_mm / neighbor_right_std_mm 为各帧左、右邻距样本标准差（字符串，mm）；
    # neighbor_*_distances_detail_mm 为按帧顺序、分号分隔的邻距明细（缺帧为「—」），主要供 Excel 列展示。
    neighbor_index_offset: int | str | None = None,
    neighbor_left_index: int | str | None = None,
    neighbor_left_distance_mm: int | str | None = None,
    neighbor_right_index: int | str | None = None,
    neighbor_right_distance_mm: int | str | None = None,
    neighbor_lr_distance_diff_mm: int | str | None = None,
    neighbor_left_std_mm: str | None = None,
    neighbor_right_std_mm: str | None = None,
    neighbor_left_distances_detail_mm: str | None = None,
    neighbor_right_distances_detail_mm: str | None = None,
) -> dict[str, Any]:
    """组装一行记录：含导出所需的 angle_deg、measured_distance、front_edge、back_edge、reflectivity。"""
    mm = int(raw["measured_distance"])
    ag = float(raw["angle_deg"])
    fe = int(raw["front_edge"])
    be = int(raw["back_edge"])
    rf = int(raw["reflectivity"])
    pm = (
        int(probe_median_index)
        if isinstance(probe_median_index, int)
        else (probe_median_index if probe_median_index is not None else "")
    )
    pmt = probe_match_target if probe_match_target is not None else ""
    pis = probe_indices_sequence if probe_indices_sequence is not None else ""
    psct = (
        probe_sequence_contains_target
        if probe_sequence_contains_target is not None
        else ""
    )
    pstc = (
        int(probe_sequence_target_hit_count)
        if isinstance(probe_sequence_target_hit_count, int)
        else (
            probe_sequence_target_hit_count
            if probe_sequence_target_hit_count is not None
            else ""
        )
    )
    pmh = (
        int(probe_max_hit_count)
        if isinstance(probe_max_hit_count, int)
        else (probe_max_hit_count if probe_max_hit_count is not None else "")
    )
    tmd = (
        int(target_minus_probe_median)
        if isinstance(target_minus_probe_median, int)
        else (target_minus_probe_median if target_minus_probe_median is not None else "")
    )
    n_io = (
        int(neighbor_index_offset)
        if isinstance(neighbor_index_offset, int)
        else (neighbor_index_offset if neighbor_index_offset is not None else "")
    )
    n_li = (
        int(neighbor_left_index)
        if isinstance(neighbor_left_index, int)
        else (neighbor_left_index if neighbor_left_index is not None else "")
    )
    n_lmm = (
        int(neighbor_left_distance_mm)
        if isinstance(neighbor_left_distance_mm, int)
        else (neighbor_left_distance_mm if neighbor_left_distance_mm is not None else "")
    )
    n_ri = (
        int(neighbor_right_index)
        if isinstance(neighbor_right_index, int)
        else (neighbor_right_index if neighbor_right_index is not None else "")
    )
    n_rmm = (
        int(neighbor_right_distance_mm)
        if isinstance(neighbor_right_distance_mm, int)
        else (neighbor_right_distance_mm if neighbor_right_distance_mm is not None else "")
    )
    n_dif = (
        int(neighbor_lr_distance_diff_mm)
        if isinstance(neighbor_lr_distance_diff_mm, int)
        else (neighbor_lr_distance_diff_mm if neighbor_lr_distance_diff_mm is not None else "")
    )
    n_lstd = "" if neighbor_left_std_mm is None else str(neighbor_left_std_mm)
    n_rstd = "" if neighbor_right_std_mm is None else str(neighbor_right_std_mm)
    n_ldet = "" if neighbor_left_distances_detail_mm is None else str(neighbor_left_distances_detail_mm)
    n_rdet = "" if neighbor_right_distances_detail_mm is None else str(neighbor_right_distances_detail_mm)
    return {
        "step": step,
        "rotate_label": rotate_label,
        "rotate_deg": float(rotate_deg),
        "target_index": int(target_index),
        "probe_median_index": pm,
        "probe_match_target": pmt,
        "probe_indices_sequence": pis,
        "probe_sequence_contains_target": psct,
        "probe_sequence_target_hit_count": pstc,
        "probe_max_hit_count": pmh,
        "target_minus_probe_median": tmd,
        "neighbor_index_offset": n_io,
        "neighbor_left_index": n_li,
        "neighbor_left_distance_mm": n_lmm,
        "neighbor_right_index": n_ri,
        "neighbor_right_distance_mm": n_rmm,
        "neighbor_lr_distance_diff_mm": n_dif,
        "neighbor_left_std_mm": n_lstd,
        "neighbor_right_std_mm": n_rstd,
        "neighbor_left_distances_detail_mm": n_ldet,
        "neighbor_right_distances_detail_mm": n_rdet,
        "index_start": index_start if index_start is not None else "",
        "index_end": index_end if index_end is not None else "",
        "delta_index": int(delta_index),
        "measured_angle_deg": float(measured_angle_deg),
        "theory_angle_deg": float(theory_angle_deg),
        "error_deg": float(error_deg),
        "distance_m": mm / 1000.0,
        "intensity": rf,
        "timestamp": timestamp,
        "radar_angle_deg": ag,
        "angle_deg": ag,
        "measured_distance": mm,
        "front_edge": fe,
        "back_edge": be,
        "reflectivity": rf,
        "anomalies": anomalies,
    }


def fmt_rotate_cumulative(deg: float) -> str:
    if abs(deg) < 1e-9:
        return "0°"
    sign = "+" if deg > 0 else "-"
    return f"{sign}{abs(deg):.1f}°"


def move_turntable(port: str, baud: int, axis: Axis, direction: Sign, pulses: int) -> None:
    cfg = Y100SCSerialConfig(port=port, baudrate=baud, timeout_s=0.5)
    with Y100SCClient(cfg) as dev:
        dev.handshake()
        dev.move(axis, direction, pulses, wait_timeout_s=180.0)


def run_newpre_sequence(args: argparse.Namespace) -> list[dict[str, Any]]:
    radar = H1CalibrationRadar(host=args.radar_ip, port=args.radar_port)
    if not radar.connect_radar():
        raise SystemExit(radar.last_error or "雷达连接失败")
    radar.configure_scan_parameters(angular_resolution_deg=args.angular_res_deg)
    radar.calibration_header_size = int(args.calibration_header_bytes)

    sign = 1 if args.dir == "+" else -1
    rows: list[dict[str, Any]] = []

    m0 = radar.optimized_single_measurement(args.index_start, args.index_end, args.max_distance_mm)
    if not m0:
        radar.close()
        raise SystemExit(radar.last_error or "step0 采集失败")
    run0, err0 = find_longest_run_for_calibration(
        m0.get("all_results") or [], args.min_distance_mm, args.max_distance_mm
    )
    if err0 or not run0:
        radar.close()
        raise SystemExit(f"step0: {err0 or '目标提取失败'}")
    s0 = snapshot_from_run(run0, args.angular_res_deg)
    init_idx = int(s0["center_index"])
    all0 = m0.get("all_results") or []
    p0 = lookup_point_by_index(all0, init_idx)
    if p0 is None:
        p0 = min(run0, key=lambda x: abs(int(x["index"]) - init_idx))
    init_mm = int(p0["measured_distance"])
    init_i = int(p0["reflectivity"])

    ts0 = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())
    rows.append(
        build_step_row(
            step=0,
            rotate_label="0°",
            rotate_deg=0.0,
            target_index=init_idx,
            index_start=int(s0["index_start"]),
            index_end=int(s0["index_end"]),
            delta_index=0,
            measured_angle_deg=0.0,
            theory_angle_deg=0.0,
            error_deg=0.0,
            timestamp=ts0,
            anomalies="",
            raw=p0,
        )
    )
    print(f"[0] 初始 center_index={init_idx}, r={init_mm}mm, I={init_i}")

    settle_s = args.settle_ms / 1000.0
    axis: Axis = args.axis  # type: ignore[assignment]
    direction: Sign = args.dir  # type: ignore[assignment]

    try:
        for k in range(1, args.steps + 1):
            move_turntable(args.com, args.baud, axis, direction, args.pulses_per_step)
            time.sleep(settle_s)

            theory_idx = init_idx + sign * k
            theory_deg = sign * k * args.step_angle_deg

            m1 = radar.optimized_single_measurement(
                args.index_start, args.index_end, args.max_distance_mm
            )
            if not m1:
                raise RuntimeError(radar.last_error or f"step{k} 采集失败")
            # step≥1：按首帧中心 index 每步 ±1，从当帧点云取该 index 的原始点（不再做邻域搜索）
            cur_idx = theory_idx
            all1 = m1.get("all_results") or []
            pk = lookup_point_by_index(all1, cur_idx)
            if pk is None:
                raise RuntimeError(f"step{k}: 点云中无 index={cur_idx}（请检查索引起止是否覆盖该 index）")
            d_idx = cur_idx - init_idx
            measured_deg = d_idx * args.angular_res_deg
            err_deg = measured_deg - theory_deg

            anomalies: list[str] = []
            dm = int(pk["measured_distance"])
            if abs(dm - init_mm) > max(500, int(0.05 * max(init_mm, 1))):
                anomalies.append("距离突变")
            di = int(pk["reflectivity"])
            if init_i > 0 and abs(di - init_i) > max(80, int(0.5 * init_i)):
                anomalies.append("强度异常")

            ts = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())
            row = build_step_row(
                step=k,
                rotate_label=fmt_rotate_cumulative(theory_deg),
                rotate_deg=float(theory_deg),
                target_index=cur_idx,
                index_start="",
                index_end="",
                delta_index=int(d_idx),
                measured_angle_deg=float(measured_deg),
                theory_angle_deg=float(theory_deg),
                error_deg=float(err_deg),
                timestamp=ts,
                anomalies="；".join(anomalies),
                raw=pk,
            )
            rows.append(row)
            extra = f" [{row['anomalies']}]" if row["anomalies"] else ""
            print(
                f"[{k}] {row['rotate_label']} idx={cur_idx} Δidx={d_idx} "
                f"测得角={measured_deg:.4f}° 理论={theory_deg:+.4f}° 误差={err_deg:+.4f}°{extra}"
            )
    finally:
        radar.close()

    return rows


def summarize(rows: list[dict[str, Any]]) -> str:
    errs = [float(r["error_deg"]) for r in rows if int(r["step"]) > 0]
    if not errs:
        return "无步进数据"
    mu = statistics.mean(errs)
    mx = max(abs(e) for e in errs)
    mn = min(errs)
    sd = statistics.stdev(errs) if len(errs) > 1 else 0.0
    lost = sum(1 for r in rows if "点位丢失" in str(r.get("anomalies", "")))
    jumps = sum(1 for r in rows if "index跳变" in str(r.get("anomalies", "")))
    steps = max(1, len(rows) - 1)
    return (
        f"平均误差 {mu:+.6f}° | 最大绝对误差 {mx:.6f}° | 最小误差 {mn:+.6f}° | 标准差 {sd:.6f}° | "
        f"丢点标记 {lost} | index跳变 {jumps} | 丢点率 {lost/steps:.2%}"
    )


def _export_row_cells(r: dict[str, Any]) -> list[Any]:
    """单行导出单元格（含每 index 原始点字段）。"""
    return [
        r["step"],
        r["rotate_label"],
        r["target_index"],
        r.get("probe_median_index", ""),
        r.get("probe_match_target", ""),
        r.get("probe_indices_sequence", ""),
        r.get("probe_sequence_contains_target", ""),
        r.get("probe_sequence_target_hit_count", ""),
        r.get("probe_max_hit_count", ""),
        r.get("target_minus_probe_median", ""),
        r["index_start"],
        r["index_end"],
        r["delta_index"],
        round(r["measured_angle_deg"], 6),
        round(r["theory_angle_deg"], 6),
        round(r["error_deg"], 6),
        round(r["distance_m"], 6),
        r["intensity"],
        r.get("neighbor_index_offset", ""),
        r.get("neighbor_left_index", ""),
        r.get("neighbor_left_distance_mm", ""),
        r.get("neighbor_left_std_mm", ""),
        r.get("neighbor_right_index", ""),
        r.get("neighbor_right_distance_mm", ""),
        r.get("neighbor_right_std_mm", ""),
        r.get("neighbor_lr_distance_diff_mm", ""),
        r.get("neighbor_left_distances_detail_mm", ""),
        r.get("neighbor_right_distances_detail_mm", ""),
        r["timestamp"],
        round(r["radar_angle_deg"], 4),
        round(float(r["angle_deg"]), 6),
        int(r["measured_distance"]),
        int(r["front_edge"]),
        int(r["back_edge"]),
        int(r["reflectivity"]),
        r.get("anomalies") or "",
    ]


def export_rows(
    rows: list[dict[str, Any]],
    out_path: Path,
    summary: str = "",
    final_conclusion: str | None = None,
) -> None:
    headers = [
        "步序",
        "转台角度",
        "目标索引",
        "距离窗最高反射中位索引",
        "中位与目标一致",
        "各次最高反射索引序列",
        "序列含目标",
        "目标在序列中次数",
        "最多命中次数",
        "目标索引减中位索引",
        "段起始索引",
        "段结束索引",
        "索引偏移",
        "雷达测得角度(°)",
        "理论角度(°)",
        "角度误差(°)",
        "距离(米)",
        "强度",
        "邻indexΔ",
        "左邻index",
        "左邻距mm",
        "左邻距σ(mm)",
        "右邻index",
        "右邻距mm",
        "右邻距σ(mm)",
        "左右距差mm",
        "左邻距各帧(mm)",
        "右邻距各帧(mm)",
        "时间戳",
        "雷达角(°)",
        "角度(°)",
        "距离(mm)",
        "前沿",
        "后沿",
        "反射率",
        "异常备注",
    ]
    suf = out_path.suffix.lower()
    if suf == ".xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
        except ImportError:
            raise SystemExit("未安装 openpyxl，请 pip install openpyxl 或改用 .csv 输出")
        wb = Workbook()
        ws = wb.active
        ws.title = "newpre"
        ws.append(headers)
        for r in rows:
            ws.append(_export_row_cells(r))
        # 邻距「各帧」列为分号分隔长文本，开启自动换行便于在 Excel 中阅读。
        _wrap_titles = ("左邻距各帧(mm)", "右邻距各帧(mm)")
        _col_wrap = {name: headers.index(name) + 1 for name in _wrap_titles if name in headers}
        for _ci in _col_wrap.values():
            for _ri in range(2, ws.max_row + 1):
                _cell = ws.cell(row=_ri, column=_ci)
                _cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.append([])
        if summary.strip():
            ws.append(["汇总", summary])
        if final_conclusion:
            ws.append(["最终结论", final_conclusion])
        wb.save(str(out_path))
    else:
        path = out_path if suf == ".csv" else out_path.with_suffix(".csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(_export_row_cells(r))
            w.writerow([])
            if summary.strip():
                w.writerow(["汇总", summary])
            if final_conclusion:
                w.writerow(["最终结论", final_conclusion])


def export_all_index_points(all_results: list[dict[str, Any]], out_path: Path) -> None:
    """
    将一帧解析后的点列表（含 index、angle_deg、measured_distance、front_edge、back_edge、reflectivity）
    按索引升序导出为 Excel 或 CSV，便于核对统计/步进所用原始点云。
    """
    headers = ["索引", "角度(°)", "距离(mm)", "前沿", "后沿", "反射率"]
    points = sorted(all_results, key=lambda p: int(p["index"]))
    suf = out_path.suffix.lower()
    if suf == ".xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            raise SystemExit("未安装 openpyxl，请 pip install openpyxl 或改用 .csv 输出")
        wb = Workbook()
        ws = wb.active
        ws.title = "索引窗点云"
        ws.append(["说明", f"共 {len(points)} 点，按索引升序，单位见表头"])
        ws.append(headers)
        for p in points:
            ws.append(
                [
                    int(p["index"]),
                    round(float(p["angle_deg"]), 6),
                    int(p["measured_distance"]),
                    int(p["front_edge"]),
                    int(p["back_edge"]),
                    int(p["reflectivity"]),
                ]
            )
        wb.save(str(out_path))
    else:
        path = out_path if suf == ".csv" else out_path.with_suffix(".csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["说明", f"共 {len(points)} 点，按索引升序"])
            w.writerow(headers)
            for p in points:
                w.writerow(
                    [
                        int(p["index"]),
                        f"{float(p['angle_deg']):.6f}",
                        int(p["measured_distance"]),
                        int(p["front_edge"]),
                        int(p["back_edge"]),
                        int(p["reflectivity"]),
                    ]
                )


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="newpre.md 连续步进角分辨率测试（CLI）")
    p.add_argument("--radar-ip", default="192.168.1.111", help="雷达 IP")
    p.add_argument("--radar-port", type=int, default=2111, help="雷达 TCP 端口")
    p.add_argument("--index-start", type=int, default=0, help="索引起")
    p.add_argument("--index-end", type=int, default=2700, help="索引止")
    p.add_argument("--min-distance-mm", type=int, default=10, help="目标距离下限（开区间）")
    p.add_argument("--max-distance-mm", type=int, default=12000, help="目标距离上限（开区间）")
    p.add_argument("--angular-res-deg", type=float, default=0.1, help="角分辨率 °/index")
    p.add_argument(
        "--calibration-header-bytes",
        type=int,
        default=0,
        help="标定 TCP 回包首点前的固定字节数；与盲区桌面脚本一致为 0，带 6 字节前缀的机型填 6",
    )
    p.add_argument("--step-angle-deg", type=float, default=0.1, help="每步理论转角（通常等于角分辨率）")
    p.add_argument("--theory-window", type=int, default=12, help="理论 index 搜索半宽")
    p.add_argument("--com", required=True, help="转台串口，如 COM6")
    p.add_argument("--baud", type=int, default=9600)
    p.add_argument("--axis", default="X", choices=["X", "Y", "Z", "r", "t", "T"])
    p.add_argument("--dir", default="+", choices=["+", "-"])
    p.add_argument(
        "--pulses-per-step",
        type=int,
        default=40,
        help="每步发送的脉冲数；与 --step-angle-deg 一致时，1°=400 脉冲则 0.1° 填 40",
    )
    p.add_argument("--settle-ms", type=int, default=300, help="转台到位后等待毫秒")
    p.add_argument("--steps", type=int, default=50, help="连续步数 N（不含 step0）")
    p.add_argument(
        "--out",
        type=Path,
        default=None,
        help="导出路径，如 result.csv 或 result.xlsx；省略则仅打印",
    )
    return p


def main() -> int:
    args = build_arg_parser().parse_args()
    if args.min_distance_mm >= args.max_distance_mm:
        print("错误: min-distance-mm 须小于 max-distance-mm", file=sys.stderr)
        return 2
    if args.index_start > args.index_end:
        print("错误: index-start 不能大于 index-end", file=sys.stderr)
        return 2

    try:
        rows = run_newpre_sequence(args)
    except (Y100SCError, OSError, RuntimeError, ValueError) as e:
        print(f"失败: {e}", file=sys.stderr)
        return 1

    summary = summarize(rows)
    print("---")
    print(summary)

    if args.out:
        export_rows(rows, args.out, summary)
        print(f"已写入: {args.out.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
