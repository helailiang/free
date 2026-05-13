#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基于 h2_resolution_gui_test / newpre 导出的连续测试表格（CSV 或 Excel），
按「步序」—「目标索引减中位索引」绘制折线图。

导出表头与 newpre_resolution_cli_test.export_rows 一致，需含列名：
  目标索引减中位索引（对应行字段 target_minus_probe_median）

用法（在「码盘补偿」目录或任意路径）:
  python analyze_h2_export_delta_median_plot.py 导出.csv
  python analyze_h2_export_delta_median_plot.py 导出.xlsx -o 分析图.png
  python analyze_h2_export_delta_median_plot.py 导出.csv --no-show

依赖: matplotlib；读 .xlsx 时需 openpyxl。
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any

COL_STEP = "步序"
COL_DELTA = "目标索引减中位索引"


def _to_int_step(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, int) and not isinstance(v, bool):
        return int(v)
    s = str(v).strip()
    if not s or s.startswith("汇总") or s.startswith("最终"):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _to_float_delta(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in ("none", "null", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _load_rows_csv(path: Path) -> list[dict[str, str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _load_rows_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("读取 .xlsx 需要 openpyxl：pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    out: list[dict[str, Any]] = []
    for tup in rows_iter:
        if tup is None or all(x is None for x in tup):
            continue
        row: dict[str, Any] = {}
        for i, name in enumerate(headers):
            if not name:
                continue
            row[name] = tup[i] if i < len(tup) else None
        if row:
            out.append(row)
    wb.close()
    return out


def load_export_rows(path: Path) -> list[dict[str, Any]]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _load_rows_csv(path)
    if suf in (".xlsx", ".xlsm"):
        return _load_rows_xlsx(path)
    raise SystemExit(f"不支持的扩展名: {suf}（请使用 .csv 或 .xlsx）")


def extract_step_and_delta(rows: list[dict[str, Any]]) -> tuple[list[int], list[float]]:
    steps: list[int] = []
    deltas: list[float] = []
    for r in rows:
        st = _to_int_step(r.get(COL_STEP))
        if st is None:
            continue
        d = _to_float_delta(r.get(COL_DELTA))
        if d is None:
            continue
        steps.append(st)
        deltas.append(d)
    pairs = sorted(zip(steps, deltas), key=lambda t: t[0])
    return [p[0] for p in pairs], [p[1] for p in pairs]


def main() -> int:
    p = argparse.ArgumentParser(description="导出表：步序 vs 目标索引减中位索引 折线图")
    p.add_argument("input", type=Path, help="GUI 导出的 .csv 或 .xlsx")
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="输出图片路径，默认与输入同目录、文件名加 _delta_median.png",
    )
    p.add_argument("--no-show", action="store_true", help="仅保存图片，不弹出交互窗口")
    args = p.parse_args()
    inp = args.input.expanduser().resolve()
    if not inp.is_file():
        print(f"文件不存在: {inp}", file=sys.stderr)
        return 1

    rows = load_export_rows(inp)
    if not rows:
        print("表格无数据行", file=sys.stderr)
        return 2

    headers = list(rows[0].keys())
    if COL_DELTA not in headers and not any(COL_DELTA in (h or "") for h in headers):
        print(f"未找到列「{COL_DELTA}」。当前表头含: {headers[:20]}…", file=sys.stderr)
        return 3
    if COL_STEP not in headers:
        print(f"未找到列「{COL_STEP}」", file=sys.stderr)
        return 3

    xs, ys = extract_step_and_delta(rows)
    if not xs:
        print(f"无有效数据点（需同时解析「{COL_STEP}」与「{COL_DELTA}」为数值）", file=sys.stderr)
        return 4

    try:
        import matplotlib.pyplot as plt
    except ImportError:
        print("需要 matplotlib：pip install matplotlib", file=sys.stderr)
        return 5

    out_path = args.output
    if out_path is None:
        out_path = inp.parent / f"{inp.stem}_delta_median.png"
    else:
        out_path = out_path.expanduser().resolve()

    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    fig, ax = plt.subplots(figsize=(10, 5), layout="constrained")
    ax.plot(xs, ys, marker="o", linewidth=1.2, markersize=4, color="#1f77b4")
    ax.set_xlabel(COL_STEP)
    ax.set_ylabel(COL_DELTA)
    ax.set_title(f"{inp.name}\n{COL_DELTA} 随步序变化")
    ax.grid(True, linestyle="--", alpha=0.35)
    ax.axhline(0, color="gray", linewidth=0.8, linestyle="-", alpha=0.6)
    fig.savefig(out_path, dpi=150)
    print(f"已保存: {out_path}")
    if not args.no_show:
        plt.show()
    else:
        plt.close(fig)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
